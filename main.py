import base64
import os
import requests
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from pdf2image import convert_from_path
import pandas as pd
from openpyxl import workbook,load_workbook
import glob

import openpyxl

API_KEY = "8yDoZvS2se42b56bk7ywoPR8"
SECRET_KEY ="5sS8gP8WIPrihF7CjnKEtoXz7sliIDAA"
if os.path.exists("差异数据.xlsx"):
    os.remove("差异数据.xlsx")




def Execute_Programme():
    print('开始执行！！！')
    # 这是你发票的存放地址，自行更改
    path ="C:\\Users\\93678\\Desktop\\TESTING"
    pics = get_pics(path)

    datas = get_datas(pics)
    save_to_excel(datas)
    file_path_list = glob.glob('C:/Users/93678/PycharmProjects/pythonProject10/*.xlsx')
    compare_multiple_excel_files(file_path_list)
    print('执行结束！')



def compare_multiple_excel_files(file_path_list):
    # 初始化一个字典，用于存储每个单元格的值
    cell_values = {}

    # 遍历文件路径列表
    for file_path in file_path_list:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 遍历每个单元格
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell_key = f"{file_path}-{cell.coordinate}"

                cell_values[cell_key] = cell.value

        wb.close()

    # 比较单元格值
    diff_data = []
    for cell_key, cell_value in cell_values.items():
        file_name, cell_coordinate = cell_key.split("-")
        row, col = openpyxl.utils.cell.coordinate_to_tuple(cell_coordinate)
        for i in range(len(file_path_list) - 1):
            next_file_path = file_path_list[i + 1]
            next_cell_key = cell_key.replace(file_path_list[0], next_file_path)
            if next_cell_key in cell_values and cell_values[cell_key] != cell_values[next_cell_key]:
                title = ws.cell(row=1, column=col).value
                print(f"Cell {cell_coordinate} in {file_name} ({next_file_path} vs. {file_path}, {row}行, {col}列, {title}): 不同的地方: {cell_value} != {cell_values[next_cell_key]}")
                diff_data.append({
                    "cell_coordinate": cell_coordinate,
                    "file_name": file_name,
                    "next_file_path": next_file_path,
                    "current_file_path": file_path,
                    "row": row,
                    "col": col,
                    "title": title,
                    "cell_value": cell_value,
                    "next_cell_value": cell_values[next_cell_key]
                })

    # 将差异数据写入新的 Excel 文件中
    if len(diff_data) > 0:
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["当前文件名", "标题 ", "当前值", "下个文件名", "标题","下一个文件的值", "行", "列"]
        for i, header in enumerate(headers, start=1):
            ws.cell(row=1, column=i, value=header)
        for i, data in enumerate(diff_data, start=2):
            ws.cell(row=i, column=1, value=data["file_name"])
            ws.cell(row=i, column=2, value=data["title"])
            ws.cell(row=i, column=3, value=data["cell_value"])
            ws.cell(row=i, column=4, value=data["next_file_path"])
            ws.cell(row=i, column=5, value=data["title"])
            ws.cell(row=i, column=6, value=data["next_cell_value"])
            ws.cell(row=i, column=7, value=data["row"])
            ws.cell(row=i, column=8, value=data["col"])

        wb.save("差异数据.xlsx")

def get_pics(path):
    print('正在生成图片路径')
    # 生成一个空列表用于存放图片路径
    pics = []
    # 检查路径是否存在
    if not os.path.exists(path):
        print('指定的路径不存在：', path)
        return pics
    # 检查路径是否为目录
    if not os.path.isdir(path):
        print('指定的路径不是一个目录：', path)
        return pics
    # 遍历目录中的文件，找到后缀为 jpg 和 png 的文件，整理之后加入列表
    for filename in os.listdir(path):

        if filename.endswith('jpg') or filename.endswith('png'):
            pic = os.path.join(path, filename)
            pics.append(pic)
        elif filename.endswith('pdf'):
            pdf=os.path.join(path,filename)
            pages=convert_from_path(pdf)
            for i,page in enumerate(pages):
                pic = os.path.join(path, f"{os.path.splitext(filename)[0]}_{i + 1}.jpg")
                page.save(pic, 'JPEG')
                pics.append(pic)
    print('图片路径生成成功！')
    return pics

def get_datas(pics):
    datas = []
    for pic in pics:
        data = get_context(pic)
        if data:
            datas.append(data)
    return datas

def get_context(pic):
    try:
        # 二进制方式打开图片文件
        with open(pic, 'rb') as f:
            img = base64.b64encode(f.read())

        url = "https://aip.baidubce.com/rest/2.0/ocr/v1/vat_invoice"
        access_token = get_access_token()
        headers = {'content-type': 'application/x-www-form-urlencoded'}
        params = {"image":img, "access_token":access_token}
        response = requests.post(url, data=params, headers=headers)

        if response.status_code == 200:
            json1 = response.json()
            if 'words_result' in json1:
                data = {}
                data['发票代码']=json1['words_result'].get( "InvoiceNumConfirm")
                data["购买方纳税人识别号"] = json1['words_result'].get("PurchaserRegisterNum", '')
                data['购买方名称'] = json1['words_result'].get('PurchaserName', '')
                data['销售方名称'] = json1['words_result'].get('SellerName', '')
                data['销售方纳税人识别号'] = json1['words_result'].get('SellerRegisterNum', '')
                data['金额(不含税)'] = json1['words_result'].get('TotalAmount', '') #金额
                data["税率"] = str(json1['words_result'].get('CommodityTaxRate'))
                data['税额']=json1['words_result'].get('TotalTax','')
                data['价税合计(含税价格)']=json1['words_result'].get('AmountInFiguers')
                data["项目名称"] = str(json1['words_result'].get("CommodityName", ''))









                return data
            else:
                print('无法识别发票内容：', pic)
        else:
            print('请求识别发票内容失败：', pic)
    except Exception as e:
        print('识别发票内容出错：', pic, e)

def get_access_token():
    """
    使用 AK，SK 生成鉴权签名（Access Token）
    :return: access_token，或是None(如果错误)
    """
    try:
        url = 'https://aip.baidubce.com/oauth/2.0/token'
        params = {"grant_type": "client_credentials", "client_id": API_KEY, "client_secret": SECRET_KEY}
        response = requests.post(url, params=params)
        if response.status_code == 200:
            access_token = response.json().get("access_token")
            return access_token
        else:
            print('获取 access_token 失败：', response.text)
    except Exception as e:
        print('获取 access_token 出错：', e)

def save_to_excel(datas):
    print('正在写入数据！')
    book = openpyxl.Workbook()
    sheet = book.active
    title = ['购买方名称', '销售方名称', '购买方纳税人识别号', '销售方纳税人识别号', '金额(不含税)','税额','价税合计(含税价格)','发票代码','项目名称']
    sheet.append(title)
    for data in datas:
        row = [data.get(field, '') for field in title]
        sheet.append(row)
    book.save('增值税发票.xlsx')
    print('数据写入成功！')



if __name__ == '__main__':
   Execute_Programme()